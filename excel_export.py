"""
==========================================
EXCEL EXPORT
AI Invoice Extractor
==========================================
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from logger import logger


# ==========================================
# HEADER STYLE
# ==========================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

CENTER = Alignment(
    horizontal="center",
    vertical="center"
)


# ==========================================
# EXPORT SINGLE DOCUMENT
# ==========================================

def export_to_excel(document: dict, output_path):

    try:

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Invoice"

        headers = [

            "Vendor Name",
            "Invoice Number",
            "Invoice Date",
            "GST Number",
            "Subtotal",
            "Tax",
            "Grand Total",
            "Payment Method",
            "Currency",
            "Filename"

        ]

        for col, header in enumerate(headers, start=1):

            cell = sheet.cell(row=1, column=col)

            cell.value = header

            cell.fill = HEADER_FILL

            cell.font = HEADER_FONT

            cell.alignment = CENTER

        values = [

            document.get("vendor_name", ""),

            document.get("invoice_number", ""),

            document.get("invoice_date", ""),

            document.get("gst_number", ""),

            document.get("subtotal", ""),

            document.get("tax", ""),

            document.get("grand_total", ""),

            document.get("payment_method", ""),

            document.get("currency", ""),

            document.get("filename", "")

        ]

        for col, value in enumerate(values, start=1):

            sheet.cell(
                row=2,
                column=col
            ).value = value

        # Auto Width

        for column in sheet.columns:

            max_length = 0

            letter = column[0].column_letter

            for cell in column:

                try:

                    max_length = max(

                        max_length,

                        len(str(cell.value))

                    )

                except:

                    pass

            sheet.column_dimensions[
                letter
            ].width = max_length + 5

        output_path = Path(output_path)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        workbook.save(output_path)

        workbook.close()

        logger.info(
            f"Excel Exported : {output_path}"
        )

        return str(output_path)

    except Exception:

        logger.exception(
            "Excel Export Failed"
        )

        raise


# ==========================================
# EXPORT MULTIPLE DOCUMENTS
# ==========================================

def export_multiple_to_excel(
    documents,
    output_path
):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Invoices"

    headers = [

        "Vendor Name",

        "Invoice Number",

        "Invoice Date",

        "GST Number",

        "Subtotal",

        "Tax",

        "Grand Total",

        "Payment Method",

        "Currency",

        "Filename"

    ]

    for col, header in enumerate(headers, start=1):

        cell = sheet.cell(row=1, column=col)

        cell.value = header

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = CENTER

    row = 2

    for document in documents:

        sheet.cell(row,1).value=document.get("vendor_name","")
        sheet.cell(row,2).value=document.get("invoice_number","")
        sheet.cell(row,3).value=document.get("invoice_date","")
        sheet.cell(row,4).value=document.get("gst_number","")
        sheet.cell(row,5).value=document.get("subtotal","")
        sheet.cell(row,6).value=document.get("tax","")
        sheet.cell(row,7).value=document.get("grand_total","")
        sheet.cell(row,8).value=document.get("payment_method","")
        sheet.cell(row,9).value=document.get("currency","")
        sheet.cell(row,10).value=document.get("filename","")

        row += 1

    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook.save(output_path)

    workbook.close()

    logger.info(
        f"Excel Exported : {output_path}"
    )

    return str(output_path)